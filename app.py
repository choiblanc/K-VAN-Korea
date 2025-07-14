import streamlit as st
import matplotlib.pyplot as plt
import pandas as pd
from datetime import datetime, timedelta
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import holidays
from io import BytesIO

try:
    plt.rc('font', family='NanumGothic')
    plt.rcParams['axes.unicode_minus'] = False
except:
    st.warning("나눔고딕 폰트가 설치되어 있지 않아 그래프의 한글이 깨질 수 있습니다. 로컬에 폰트를 설치하거나 Streamlit 클라우드에 폰트 파일을 추가하세요.")
    pass

kr_holidays = holidays.KR(years=2025)

base_minutes = {
    '부강연동 1A': 6*60+20, '부강연동 2A': 7*60+20, '부강연동 3B': 8*60+50, '부강연동 4B': 8*60+50,
    '부강연동 5C': 8*60+10, '부강연동 6C': 8*60+50, '부강연동 7D': 8*60, '전동주중 1A': 9*60+40,
    '전동주중 2B': 9*60+40, '전동주중 3C': 9*60+40, '부강연동 1A2A': 8*60, '부강연동 3B4B': 8*60,
    '부강연동 5C6C': 8*60, '전동 1A2B': 9*60, '전동 3C4D': 9*60, '고정1': 10*60, '고정2': 10*60,
    '고정3(오전)': 7*60+30, '고정3(오후)': 8*60+10, '당직': 7*60, '휴무': 0
}

weekday_add = {
    '부강연동 7D': 20, '부강연동 5C6C': 5, '고정1': 5, '고정3(오전)': 15, '고정3(오후)': 30, '당직': 30,
}

weekend_add = {
    '부강연동 5C6C': 10, '고정1': 10, '고정3(오전)': 30, '고정3(오후)': 60, '당직': 60,
}

weekend_15x = [
    '부강연동 1A2A', '부강연동 3B4B', '부강연동 5C6C', '전동 1A2B', '전동 3C4D',
    '고정1', '고정2', '고정3(오전)', '고정3(오후)'
]

base_pattern = [
    '당직', '휴무', '고정3(오후)', '부강연동 7D', '부강연동 1A', '전동주중 1A', '고정1',
    '휴무', '부강연동 2A', '부강연동 3B', '전동주중 2B', '고정2', '부강연동 4B',
    '휴무', '전동주중 3C', '부강연동 5C', '부강연동 6C', '고정3(오전)'
]
saturday_pattern = [
    '휴무', '부강연동 1A2A', '휴무', '부강연동 5C6C', '고정3(오전)', '당직', '휴무', '고정3(오후)', '휴무', '전동 1A2B',
    '전동 3C4D', '고정1', '휴무', '부강연동 3B4B', '휴무', '휴무', '고정2', '휴무'
]
sunday_pattern = [
    '휴무', '휴무', '부강연동 1A2A', '휴무', '휴무', '고정3(오전)', '당직', '휴무', '고정3(오후)', '휴무', '전동 1A2B',
    '전동 3C4D', '고정1', '휴무', '부강연동 3B4B', '부강연동 5C6C', '휴무', '고정2'
]

patterns = {'weekday': base_pattern, 'saturday': saturday_pattern, 'sunday': sunday_pattern}
# drivers = [f"driver{i+1}" for i in range(18)]

holiday_dates = set(kr_holidays.keys())
fixed_roles = ['고정1', '고정2', '고정3(오전)', '고정3(오후)']

vis_data = [
    {"area": "부강연동 1A", "start": "06:20", "end": "13:30"}, {"area": "부강연동 2A", "start": "06:20", "end": "15:00"},
    {"area": "부강연동 3B", "start": "07:30", "end": "16:20"}, {"area": "부강연동 4B", "start": "07:30", "end": "16:20"},
    {"area": "부강연동 5C", "start": "07:40", "end": "15:50"}, {"area": "부강연동 6C", "start": "07:30", "end": "16:20"},
    {"area": "부강연동 7D", "start": "08:10", "end": "16:10"}, {"area": "전동주중 1A", "start": "09:40", "end": "19:20"},
    {"area": "전동주중 2B", "start": "09:40", "end": "19:20"}, {"area": "전동주중 3C", "start": "09:40", "end": "19:20"},
    {"area": "고정1", "start": "07:00", "end": "17:00"}, {"area": "고정2", "start": "07:00", "end": "17:00"},
    {"area": "고정3(오전)", "start": "07:00", "end": "14:30"}, {"area": "고정3(오후)", "start": "08:10", "end": "16:10"},
    {"area": "당직", "start": "05:00", "end": "13:00"}, {"area": "휴무", "start": "00:00", "end": "00:00"},
]

weekend_time_info = {
    "부강연동 1A2A": ("07:00", "15:00"), "부강연동 3B4B": ("07:40", "15:30"), "부강연동 5C6C": ("07:00", "15:00"),
    "전동 1A2B": ("09:40", "18:00"), "전동 3C4D": ("09:40", "18:00"), "고정1": ("07:00", "16:00"),
    "고정2": ("07:00", "16:00"), "고정3(오전)": ("07:00", "14:30"), "고정3(오후)": ("08:10", "16:20"),
}

dual_vehicle_weekend_areas = {
    "부강연동 5C6C": [("07:00", "11:00"), ("11:30", "15:00")],
    "전동 1A2B": [("09:40", "13:40"), ("14:10", "18:00")],
}

def is_holiday(date):
    return date in holiday_dates

def get_pattern_type(date):
    if date.weekday() == 5: return 'saturday'
    elif date.weekday() == 6: return 'sunday'
    elif is_holiday(date):
        weekday = date.weekday()
        to_saturday = (5 - weekday) % 7
        to_sunday = (6 - weekday) % 7
        return 'sunday' if to_saturday < to_sunday else 'saturday'
    else: return 'weekday'

def find_fixed_index(pattern, fixed_role_name):
    for i, duty in enumerate(pattern):
        if duty == fixed_role_name: return i
    return None

def shift_pattern(pattern, shift_amount):
    return pattern[-shift_amount:] + pattern[:-shift_amount]

def get_majority_shift(today_pattern, reference_pattern, fixed_roles):
    shifts = []
    for role in fixed_roles:
        ref_idx = find_fixed_index(reference_pattern, role)
        today_idx = find_fixed_index(today_pattern, role)
        if ref_idx is not None and today_idx is not None:
            shift = (ref_idx - today_idx) % len(today_pattern)
            shifts.append(shift)
    if not shifts: return 0
    return max(set(shifts), key=shifts.count)

def make_schedule_with_majority_shift(patterns, drivers, start_date, days=28):
    num_drivers = len(drivers)
    schedule = {driver: [] for driver in drivers}  # 모든 드라이버 포함 보장
    date_list = [start_date + timedelta(days=i) for i in range(days)]
    driver_indices = list(range(num_drivers))
    reference_pattern = None

    for day in range(days):
        today = date_list[day]
        pattern_type = get_pattern_type(today)
        daily_pattern = patterns[pattern_type][:]  # 복사본 생성

        # 드라이버 수에 맞게 pattern 확장 (휴무로 채움)
        if len(daily_pattern) < num_drivers:
            daily_pattern += ['휴무'] * (num_drivers - len(daily_pattern))
        elif len(daily_pattern) > num_drivers:
            daily_pattern = daily_pattern[:num_drivers]

        if reference_pattern is None:
            reference_pattern = daily_pattern
            shifted_pattern = daily_pattern
        else:
            shift_amount = get_majority_shift(daily_pattern, reference_pattern, fixed_roles)
            shifted_pattern = shift_pattern(daily_pattern, shift_amount)

        for i, driver in enumerate(drivers):
            duty_idx = driver_indices[i]
            schedule[driver].append(shifted_pattern[duty_idx])

        driver_indices = [(idx - 1) % num_drivers for idx in driver_indices]

    return schedule, date_list


def schedule_to_dataframe(schedule, drivers, date_list):
    columns = [d.strftime("%Y-%m-%d") for d in date_list]
    data = [schedule[driver] for driver in drivers]
    return pd.DataFrame(data, index=drivers, columns=columns)

def count_work_by_driver_and_area(schedule, date_list):
    def is_weekend(date): return date.weekday() in (5, 6) or is_holiday(date)
    driver_area_stats = defaultdict(lambda: defaultdict(lambda: {'평일근무':0, '주말근무':0}))
    for driver, duties in schedule.items():
        for duty, date in zip(duties, date_list):
            if '휴무' not in str(duty):
                area = str(duty).strip()
                if is_weekend(date): driver_area_stats[driver][area]['주말근무'] += 1
                else: driver_area_stats[driver][area]['평일근무'] += 1
    return driver_area_stats

def create_pivot_work_df(stats):
    rows = []
    for driver, area_dict in stats.items():
        for area, counts in area_dict.items():
            rows.append({'운전자': driver, '구역': area, '근무횟수': counts['평일근무'] + counts['주말근무'], '평일근무': counts['평일근무'], '주말근무': counts['주말근무']})
    df = pd.DataFrame(rows)
    if df.empty: return pd.DataFrame()
    pivot_total = pd.pivot_table(df, values='근무횟수', index='구역', columns='운전자', fill_value=0, aggfunc='sum')
    pivot_weekday = pd.pivot_table(df, values='평일근무', index='구역', columns='운전자', fill_value=0, aggfunc='sum')
    pivot_weekend = pd.pivot_table(df, values='주말근무', index='구역', columns='운전자', fill_value=0, aggfunc='sum')
    pivot_df = pd.concat({'총근무': pivot_total, '평일근무': pivot_weekday, '주말근무': pivot_weekend}, axis=1)
    driver_totals = pd.DataFrame({'총근무': pivot_total.sum(axis=0), '평일근무': pivot_weekday.sum(axis=0), '주말근무': pivot_weekend.sum(axis=0)}).T
    driver_totals.index = pd.MultiIndex.from_product([['합계'], ['총근무', '평일근무', '주말근무']])
    final_pivot = pd.concat([pivot_df, driver_totals], axis=1)
    return final_pivot

def get_work_minutes(area, date):
    base = base_minutes.get(area, 0)
    if base == 0: return 0
    weekend = date.weekday() >= 5 or is_holiday(date)
    if weekend:
        add = weekend_add.get(area, 0)
        return int(base * 1.5 + add) if area in weekend_15x else base + add
    else:
        add = weekday_add.get(area, 0)
        return base + add

def calc_monthly_and_yearly_pay(schedule, date_list, hourly_wage=12036):
    pay_data = defaultdict(lambda: defaultdict(int))
    for driver, duties in schedule.items():
        for duty, date in zip(duties, date_list):
            minutes = get_work_minutes(duty, date)
            month = date.strftime('%Y-%m')
            pay_data[driver][month] += minutes
    df = pd.DataFrame(pay_data).T.fillna(0).astype(int)
    pay_df = df.applymap(lambda x: round(x/60 * hourly_wage))
    if not pay_df.empty:
        pay_df['연간합계'] = pay_df.sum(axis=1)
        total_row = pay_df.sum(axis=0)
        total_row.name = '전체합계'
        pay_df = pd.concat([pay_df, total_row.to_frame().T])
        month_cols = sorted([c for c in pay_df.columns if c not in ['연간합계']])
        pay_df = pay_df[month_cols + ['연간합계']]
    return pay_df

def time_to_minutes(t_str):
    h, m = map(int, t_str.split(":"))
    return h * 60 + m

def create_schedule_visualization(schedule, date_list, vis_data, drivers, target_date_str):
    target_date = datetime.strptime(target_date_str, "%Y-%m-%d").date()
    try:
        date_idx = date_list.index(target_date) 
    except ValueError:
        st.error(f"{target_date_str}는 생성된 스케줄 범위에 없습니다.")
        return None

    # 안전한 duties 리스트 생성
    duties_on_date = []
    missing_drivers = []
    for driver in drivers:
        if driver in schedule and date_idx < len(schedule[driver]):
            duties_on_date.append(schedule[driver][date_idx])
        else:
            duties_on_date.append("휴무")  # 또는 "데이터 없음"
            missing_drivers.append(driver)

    if missing_drivers:
        st.warning(f"다음 드라이버는 스케줄에 없어서 기본값으로 대체됨: {', '.join(missing_drivers)}")

    pattern_type = get_pattern_type(target_date)

    time_info_weekday = {
        d['area']: (
            time_to_minutes(d['start']),
            time_to_minutes(d['end']) - time_to_minutes(d['start']),
            d['start'], d['end']
        ) for d in vis_data
    }

    time_info_weekend = {
        k: (
            time_to_minutes(v[0]),
            time_to_minutes(v[1]) - time_to_minutes(v[0]),
            v[0], v[1]
        ) for k, v in weekend_time_info.items()
    }

    time_info = time_info_weekday if pattern_type == 'weekday' else {**time_info_weekday, **time_info_weekend}

    fig, ax = plt.subplots(figsize=(12, len(drivers) * 0.4 + 2))
    y_pos = range(len(drivers))

    for i, (driver, duty) in enumerate(zip(drivers, duties_on_date)):
        if duty in dual_vehicle_weekend_areas and pattern_type in ['saturday', 'sunday']:
            shifts = dual_vehicle_weekend_areas[duty]
            colors = ['lightblue', 'deepskyblue']
            for shift_idx, (start_str, end_str) in enumerate(shifts):
                start_min, end_min = time_to_minutes(start_str), time_to_minutes(end_str)
                duration = end_min - start_min
                ax.barh(i, duration, left=start_min, height=0.6, color=colors[shift_idx], edgecolor='black')
                ax.text(start_min + 5, i, f"{duty} ({shift_idx+1}부)", va='center', ha='left', fontsize=8, color='black')
        else:
            start_min, duration, start_str, end_str = time_info.get(duty, (0, 0, '00:00', '00:00'))
            if duration > 0:
                ax.barh(i, duration, left=start_min, height=0.6, color="skyblue", edgecolor='black')
                ax.text(start_min + duration / 2, i, f"{duty}\n{start_str}~{end_str}", va='center', ha='center', fontsize=8, color='black')

                # 휴식시간 표시
                rest_start = start_min + 180  # 3시간 뒤부터 휴식 시작
                while rest_start + 30 <= start_min + duration:
                    ax.barh(i, 30, left=rest_start, height=0.6, color="#4682B4", alpha=0.8)
                    rest_start += 210  # 휴식 30분 + 3시간(210분 간격)

    ax.set_yticks(y_pos)
    ax.set_yticklabels(drivers)
    ax.invert_yaxis()
    ax.set_xlim(290, 1200)
    ax.set_xticks(range(300, 1201, 60))
    ax.set_xticklabels([f"{h:02d}:00" for h in range(5, 21)])
    ax.set_xlabel("시간")
    kor_weekdays = ['월', '화', '수', '목', '금', '토', '일']
    weekday_name_kr = kor_weekdays[target_date.weekday()]
    ax.set_title(f"{target_date_str} ({weekday_name_kr}) 드라이버별 근무 시간", pad=20)
    plt.grid(axis='x', linestyle='--', alpha=0.6)
    plt.tight_layout()
    return fig


def color_schedule(val):
    color_map = {
        "당직": "#FFA500", "휴무": "#FFFFFF", "고정3(오후)": "#FF0000", "고정3(오전)": "#FF0000",
        "고정1": "#FF0000", "고정2": "#FF0000", "부강연동 1A": "#FFFF00", "부강연동 2A": "#FFFF00",
        "부강연동 3B": "#FFFF00", "부강연동 4B": "#FFFF00", "부강연동 5C": "#FFFF00",
        "부강연동 6C": "#FFFF00", "부강연동 7D": "#FFFF00", "부강연동 1A2A": "#FFDAB9",
        "부강연동 3B4B": "#FFDAB9", "부강연동 5C6C": "#FFDAB9", "전동주중 1A": "#00B050",
        "전동주중 2B": "#00B050", "전동주중 3C": "#00B050", "전동 1A2B": "#87CEEB", "전동 3C4D": "#87CEEB",
    }
    color = color_map.get(val, "#FFFFFF")
    return f'background-color: {color}; color: black'

# ====== 엑셀 변환 함수 ======
@st.cache_data
def to_excel(schedule_df, pivot_df, pay_df):
    from openpyxl.styles import PatternFill

    color_map = {
        "당직": "FFA500", "휴무": "FFFFFF", "고정3(오후)": "FF0000", "고정3(오전)": "FF0000",
        "고정1": "FF0000", "고정2": "FF0000", "부강연동 1A": "FFFF00", "부강연동 2A": "FFFF00",
        "부강연동 3B": "FFFF00", "부강연동 4B": "FFFF00", "부강연동 5C": "FFFF00",
        "부강연동 6C": "FFFF00", "부강연동 7D": "FFFF00", "부강연동 1A2A": "FFDAB9",
        "부강연동 3B4B": "FFDAB9", "부강연동 5C6C": "FFDAB9", "전동주중 1A": "00B050",
        "전동주중 2B": "00B050", "전동주중 3C": "00B050", "전동 1A2B": "87CEEB", "전동 3C4D": "87CEEB",
    }

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        schedule_df.to_excel(writer, sheet_name="1_근무스케줄")
        pivot_df.to_excel(writer, sheet_name="2_구역별근무요약")
        pay_df.to_excel(writer, sheet_name="3_월별연간급여")

        workbook = writer.book

        ws1 = writer.sheets["1_근무스케줄"]
        for row_idx, driver in enumerate(schedule_df.index, start=2):
            for col_idx, date in enumerate(schedule_df.columns, start=2):
                value = schedule_df.loc[driver, date]
                hex_color = color_map.get(value, None)
                if hex_color:
                    fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
                    ws1.cell(row=row_idx, column=col_idx).fill = fill

        ws3 = writer.sheets["3_월별연간급여"]
        currency_format = '#,##0"원"'
        for row in range(2, 2 + pay_df.shape[0]):
            for col in range(2, 2 + pay_df.shape[1]):
                cell = ws3.cell(row=row, column=col)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = currency_format
        
        for col_idx, column_cells in enumerate(ws3.columns, start=1):
            max_length = 0
            for cell in column_cells:
                try:
                    cell_value = str(cell.value)
                    if cell.number_format == currency_format:
                        cell_value += "원"
                    max_length = max(max_length, len(cell_value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws3.column_dimensions[chr(64 + col_idx)].width = adjusted_width


    return output.getvalue()



st.set_page_config(page_title="근무 스케줄 생성기", layout="wide")
st.title("근무 스케줄 자동 생성 및 분석")

# 사이드바: 사용자 입력
with st.sidebar.expander("👤 드라이버 정보 입력 (18명)"):
    driver_infos = []
    for i in range(18):
        st.markdown(f"**[{i+1}]번째 드라이버 정보**")
        name = st.text_input(f"이름_{i}", key=f"name_{i}")
        birth = st.date_input(f"생년월일_{i}", value=datetime(1990,1,1), key=f"birth_{i}")
        gender = st.selectbox(f"성별_{i}", options=["남", "여"], key=f"gender_{i}")
        phone = st.text_input(f"전화번호_{i}", key=f"phone_{i}")
        driver_infos.append({
            "name": name if name else f"driver{i+1}",
            "birth": birth.strftime("%Y-%m-%d"),
            "gender": gender,
            "phone": phone
        })

drivers = [info["name"] for info in driver_infos]

with st.sidebar:
    st.header("설정")
    start_date = st.date_input("시작 날짜", datetime(2025, 1, 1))
    days = st.slider("생성할 기간 (일)", 1, 365, 30) # 기본 30일
    hourly_wage = st.number_input("시간당 임금 (원)", min_value=0, value=12036, step=100)

    if st.button("스케줄 생성 및 분석 실행", type="primary"):
        with st.spinner('스케줄을 생성하고 분석하는 중입니다... 잠시만 기다려주세요.'):
            schedule, date_list = make_schedule_with_majority_shift(patterns, drivers, start_date, days)

            schedule_df = schedule_to_dataframe(schedule, drivers, date_list)

            stats = count_work_by_driver_and_area(schedule, date_list)
            pivot_df = create_pivot_work_df(stats)

            pay_df = calc_monthly_and_yearly_pay(schedule, date_list, hourly_wage)

            st.session_state['schedule'] = schedule
            st.session_state['date_list'] = date_list
            st.session_state['schedule_df'] = schedule_df
            st.session_state['pivot_df'] = pivot_df
            st.session_state['pay_df'] = pay_df
            st.session_state['vis_data'] = vis_data
        
        st.success("생성이 완료되었습니다! 아래 탭에서 결과를 확인하세요.")

if 'schedule_df' not in st.session_state:
    st.info("왼쪽 사이드바에서 설정을 마친 후 '스케줄 생성' 버튼을 눌러주세요.")
else:
    schedule = st.session_state['schedule']
    date_list = st.session_state['date_list']
    schedule_df = st.session_state['schedule_df']
    pivot_df = st.session_state['pivot_df']
    pay_df = st.session_state['pay_df']
    vis_data_main = st.session_state['vis_data']

    excel_data = to_excel(schedule_df, pivot_df, pay_df)

    st.download_button(
        label="전체 데이터를 엑셀 파일로 다운로드",
        data=excel_data,
        file_name=f"Schedule_{start_date.strftime('%Y%m%d')}_{days}일.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    
    tab1, tab2, tab3, tab4 = st.tabs(["근무 스케줄", "구역별 근무 요약", "월별/연간 급여", "일일 근무 시각화"])

    with tab1:
        st.header("전체 근무 스케줄")
        styled_df = schedule_df.style.applymap(color_schedule)
        st.dataframe(styled_df, use_container_width=True, height=680)
        # st.dataframe(schedule_df)

    with tab2:
        st.header("구역별 근무 횟수 요약")
        st.dataframe(pivot_df.style.format(precision=0))

    with tab3:
        st.header(f"월별/연간 급여 (시급: {hourly_wage:,}원 기준)")
        styled_pay_df = pay_df.style.format(lambda x: f"{x:,.0f}원")
        st.dataframe(styled_pay_df, height=700)

    with tab4:
        st.header("일일 근무 시간 시각화")
        date_list_str = [d.strftime("%Y-%m-%d") for d in date_list]
        
        target_date_str = st.selectbox(
            "시각화할 날짜를 선택하세요:",
            options=date_list_str,
            index=len(date_list_str) - 1 if "고정3(오후)" in saturday_pattern else 0 # 초기 선택
        )
        
        if target_date_str:
            fig = create_schedule_visualization(schedule, date_list, vis_data_main, drivers, target_date_str)
            if fig:
                st.pyplot(fig)